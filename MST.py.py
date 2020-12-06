import argparse
import socket
import os
import time
import extractDataFromImages2
import Xtract_imgAndmails
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import email,ssl
import getpass

if __name__ == '__main__':
    
    description = """ Multi purpose security tool V 1.0  :))
    Executing opcions
    
    1) Searching emails directions of an organization:
    MST.py -op 1 -organization "name of the organization"

    2) Analizing metadata from images:
    MST.py -op 2 -path "path of the images"

    3)Get images and email addresses from a website:
    MST.py -op 3 -url url of the website

    4)Obtain server banner form a website:
    MST.py -op 4 -target " website url " -port "port to conect"

    5)Send mass mails with a list of directions:
    MST.py -op 5 -from "from mail direction" -to "path or filename of directions"
    -subject "subject of the mail" -msg "message to send"
    



                                                """
    parser = argparse.ArgumentParser(description= "Multi purpose security tool:)",epilog=
                                      description,formatter_class=argparse
                                     .RawDescriptionHelpFormatter)
    parser.add_argument("-op",metavar="OP",dest="op",help="select an executing option",
                        required=True)
    parser.add_argument("-organization",metavar="ORGA",dest="orga",help="name of the organization")
    parser.add_argument("-path",metavar="PATH",dest="path",help="path of the images")
    parser.add_argument("-url",metavar="URL",dest="url",help="Set the url to get the info")
    parser.add_argument("-target",metavar="TARGET",dest="target",help="Set the url to get the banner")
    parser.add_argument("-port",metavar="PORT",dest="port",help="port for conection to the web server")
    parser.add_argument("-from",metavar="FROM",dest="From",help="set the sender account")
    parser.add_argument("-to",metavar="TO",dest="filename",help="path to the file with the directions")
    parser.add_argument("-subject",metavar="SUBJECT",dest="asunto",help="set the Subject of the mail")
    parser.add_argument("-msg",metavar="MSG",dest="messg",help="set the message to send")
    
    params=parser.parse_args()


    time.sleep(2)
    print("""Call trans opt: received. 2-19-98 13:24:18 REC:Loc

         Trace program: running

               wake up, Neo...
            the matrix has you
          follow the white rabbit.

              knock, knock, Neo.

                            (`.         ,-,
                            ` `.    ,;' /
                             `.  ,'/ .'
                              `. X /.'
                    .-;--''--.._` ` (
                  .'            /   `
                 ,           ` '   Q '
                 ,         ,   `._    \
              ,.|         '     `-.;_'
              :  . `  ;    `  ` --,.._;
               ' `    ,   )   .'
                  `._ ,  '   /_
                     ; ,''-,;' ``-
                      ``-..__``--`

    """)
    time.sleep(2)
    print("Template by Rapid7")
    print("\n")
    print("""          
              ###########################################                                          
              #    Multi purpose security tool V 1.0    #
              ###########################################

              Created by: Eric Valenzuela 
    """)

    print( """
                Alright, partner
                Keep on rollin´! you know what time it is...
                
                        Happy hunting :))

                    """)
                           
    time.sleep(1)

    choice = params.op
    if choice == str(1):
        print("Executing Hunter for the email directions search...")
        print("Searching directions for " + params.orga)
        time.sleep(2)
    # Hunter API SEARCH:

        from pyhunter import PyHunter
        from openpyxl import Workbook
        hunter = PyHunter('5245a75326284396c2f1d962a7f1f9aaee038894')

        def Busqueda(organizacion):
            busqueda = 1
            resultado = hunter.domain_search(company=organizacion,
                                             limit=busqueda,
                                             emails_type='personal')
            return resultado

        #Guardando los resultados de la busqueda


        def GuardarInformacion(datosEncontrados, organizacion):
            libro = Workbook()
            hoja = libro.create_sheet(organizacion)
            libro.save("Hunter" + organizacion + ".xlsx")
            hoja.cell(1, 1, "Dominio")
            hoja.cell(1, 2, datosEncontrados['domain'])
            hoja.cell(2, 1, "Organización")
            hoja.cell(2, 2, datosEncontrados['organization'])
            hoja.cell(3, 1, "Correo")
            hoja.cell(3, 2, datosEncontrados['emails'][0]['value'])
            hoja.cell(4, 1, "Nombre(s)")
            hoja.cell(4, 2, datosEncontrados['emails'][0]['first_name'])
            hoja.cell(5, 1, "Apellidos")
            hoja.cell(5, 2, datosEncontrados['emails'][0]['last_name'])
            libro.save("Hunter" + organizacion + ".xlsx")


        print("Searching...")
        orga = params.orga
        datosEncontrados = Busqueda(orga)
        if datosEncontrados is None:
            print("No data found")
            exit()
        else:
            print("Got it !")
            time.sleep(2)
            print(datosEncontrados)
            print(type(datosEncontrados))
            GuardarInformacion(datosEncontrados, orga)
            print("Data saved in xlsx file")
            print(" see you next time:)")
            exit()


                  

    if choice == str(2):
        print("Extracting data from images...")
        time.sleep(2)
        try:
            extractDataFromImages2.printMeta(params.path)
            print("Done")
            print("See you next time :) MST")
            exit()
        except ModuleNotFoundError:
            print(" Module PILLOW is not installed.")
            print("Insralling PILLOW...")
            time.sleep(2)
            os.system("pip install pillow")
            print('Done.')
            time.sleep(2)
            print("Extracting data from images...")
            extractDataFromImages2.printMeta(params.path)
            print("Done")
            print("See you next time :) MST")
            exit()
            
    

    if choice == str(3):
        Xtract_imgAndmails.obtenerDatos(params.url)
        print("buscando...")
        print(" Done.")
        print("images saved in imagenes directory.")
        exit()

    if choice == str(4):
        print("Searching Banner...")
        sock=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        sock.connect((params.target,int(params.port)))
        sock.settimeout(2)
        
        targetBytes=params.target.encode()
        httpGet = b"GET /index.html HTTP/1.1\r\nHost: "+targetBytes+b"\r\nAccept: */*\r\n\r\n"
        data = ""

        try:
            sock.sendall(httpGet)
            data = sock.recvfrom(1024)
            print(data[0].decode())
            results=open("banner.txt","w")
            results.write(data[0].decode())
            results.close()
            print("Info saved in file banner.txt")
        except socket.error:
            print("Socket error",socket.errno)
        finally:
            print("closing connection...")
            sock.close()
        
    if choice == str(5):
        print("Sending mails ...")
        user = params.From
        passwd = getpass.getpass()

        message = params.messg
        file_dirs = params.filename
        destinatarios = []
        direcciones = open(file_dirs,"r")
        lines = direcciones.readlines()
        for line in lines:
            destinatarios.append(line)
         
        for destinatario in destinatarios:
            msg = MIMEMultipart()
            msg['To'] = destinatario
            msg['From'] = user
            msg['Subject'] = params.asunto

            msg.attach(MIMEText(message, 'plain'))
            
            msg.attach(MIMEText(message, 'plain'))

            context = ssl.create_default_context()
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                server.login(user, passwd)
                server.sendmail(user, destinatario, msg.as_string())
                print("successfully sent email to:",destinatario)
        print("Done")


     
          
            


        




